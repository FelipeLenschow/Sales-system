#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
import ctypes
import platform
import requests
import uuid
import qrcode
from PIL import Image, ImageTk  # Ensure both Image and ImageTk are imported
import threading
import time
import numpy as np
import math

import config


# Constants for UI scaling
BASE_WIDTH = 1920
BASE_HEIGHT = 1080
Version = "0.3.1"

def is_numlock_on():
    if platform.system() != 'Windows':
        return True  # Assume Num Lock is on for non-Windows systems
    hllDll = ctypes.WinDLL ("User32.dll")
    VK_NUMLOCK = 0x90
    return hllDll.GetKeyState(VK_NUMLOCK) & 1


def set_numlock(state=True):

    if platform.system() != 'Windows':
        return  # Do nothing for non-Windows systems

    hllDll = ctypes.WinDLL ("User32.dll")
    VK_NUMLOCK = 0x90
    KEYEVENTF_EXTENDEDKEY = 0x0001
    KEYEVENTF_KEYUP = 0x0002

    current_state = is_numlock_on()
    if current_state != state:
        # Simulate key press
        hllDll.keybd_event(VK_NUMLOCK, 0x45, KEYEVENTF_EXTENDEDKEY | 0, 0)
        hllDll.keybd_event(VK_NUMLOCK, 0x45, KEYEVENTF_EXTENDEDKEY | KEYEVENTF_KEYUP, 0)

class ProductDatabase:
    def __init__(self, filepath='produtos.xlsx'):
        self.filepath = filepath
        self.load_products()

    def load_products(self):
        try:
            # Tentar carregar o arquivo com MultiIndex no cabeçalho (2 linhas)
            self.df = pd.read_excel(self.filepath, header=[0, 1], dtype=str)

            # **Validação: Verificar se o DataFrame está vazio**
            if self.df.empty:
                raise ValueError("O arquivo está vazio.")

            # **Validação: Garantir que o cabeçalho tem o formato MultiIndex esperado**
            if not isinstance(self.df.columns, pd.MultiIndex):
                raise ValueError("O arquivo não possui um cabeçalho MultiIndex com duas linhas.")

            # Limpar espaços em branco nos cabeçalhos
            self.df.columns = pd.MultiIndex.from_tuples(
                [(str(x[0]).strip(), str(x[1]).strip()) for x in self.df.columns]
            )

            # Identificar as lojas no nível 0 do MultiIndex (exceto 'Todas')
            self.shops = [shop for shop in self.df.columns.levels[0] if shop != 'Todas']

            # Converter tipos das colunas
            self.df[('Todas', 'Codigo de Barras')] = self.df[('Todas', 'Codigo de Barras')].astype(str)
            self.df[('Metadata', 'Excel Row')] = self.df.index + 3
            for shop in self.shops:
                self.df[(shop, 'Preco')] = pd.to_numeric(self.df[(shop, 'Preco')], errors='coerce')
                self.df[(shop, 'Promo Preco')] = pd.to_numeric(self.df[(shop, 'Promo Preco')], errors='coerce')
                self.df[(shop, 'Promo Quantidade')] = pd.to_numeric(self.df[(shop, 'Promo Quantidade')],
                                                                    errors='coerce')

        except FileNotFoundError:
            messagebox.showerror("Erro", f"Arquivo {self.filepath} não encontrado.")
            self.df = pd.DataFrame()
            self.shops = []  # Inicializa como lista vazia para evitar novos erros
        except ValueError as ve:
            messagebox.showerror("Erro", f"Falha ao carregar produtos: {ve}")
            self.df = pd.DataFrame()
            self.shops = []
        except Exception as e:
            messagebox.showerror("Erro", f"Erro inesperado: {e}")
            self.df = pd.DataFrame()
            self.shops = []

    def add_product(self, product_info, shop):
        try:
            wb = load_workbook(self.filepath)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = "Produtos"
            headers = [
                ("Todas", "Codigo de Barras"),
                ("Todas", "Sabor"),
                ("Todas", "Categoria"),
                (shop, "Preco"),
                (shop, "Promo Preco"),
                (shop, "Promo Quantidade")
            ]
            for col, (header1, header2) in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header1)
                ws.cell(row=2, column=col, value=header2)

        # Map headers to column indices
        header_map = {
            f"{ws.cell(row=1, column=col).value} {ws.cell(row=2, column=col).value}": col
            for col in range(1, ws.max_column + 1)
        }
        excel_row = product_info.get('indexExcel', None)
        if excel_row is None:
            # Encontrar a próxima linha vazia para o código de barras
            data_start_row = 3
            barcode_col = header_map.get("Todas Codigo de Barras")
            excel_row = data_start_row
            while ws.cell(row=excel_row, column=barcode_col).value:
                excel_row += 1

        # Adicionar os dados na mesma linha
        ws.cell(row=excel_row, column=header_map["Todas Codigo de Barras"], value=product_info['barcode'])
        ws.cell(row=excel_row, column=header_map["Todas Sabor"], value=product_info['sabor'])
        ws.cell(row=excel_row, column=header_map["Todas Categoria"], value=product_info['categoria'])

        # Adicionar Preço
        preco = product_info['preco']
        ws.cell(row=excel_row, column=header_map[f"{shop} Preco"], value=float(preco))

        # Adicionar Promo Preço, se disponível
        promo_preco = product_info.get('promo_preco', None)
        ws.cell(row=excel_row, column=header_map[f"{shop} Promo Preco"],
                value=float(promo_preco) if promo_preco is not None else "")

        # Adicionar Promo Quantidade, se disponível
        promo_qt = product_info.get('promo_qt', None)
        ws.cell(row=excel_row, column=header_map[f"{shop} Promo Quantidade"],
                value=int(promo_qt) if promo_qt is not None else "")

        # Salvar e recarregar produtos
        wb.save(self.filepath)
        self.load_products()

    def filter_products(self, search_term, shop):
        if self.df.empty:
            return pd.DataFrame()

        df_copy = self.df.copy()
        df_copy[(shop, 'Preco')] = pd.to_numeric(df_copy[(shop, 'Preco')], errors='coerce')
        df_copy[(shop, 'Preco')] = df_copy[(shop, 'Preco')].apply(
            lambda x: f"{x:.2f}".replace('.', ',') if pd.notnull(x) else "" )

        mask = (
                self.df['Todas', 'Categoria'].str.contains(search_term, case=False, na=False) |
                self.df['Todas', 'Sabor'].str.contains(search_term, case=False, na=False) |
                df_copy[(shop, 'Preco')].astype(str).str.contains(search_term, case=False, na=False)
        )
        return self.df[mask]

    def get_unique_values(self, column, shop=None):
        if shop:
            return sorted(self.df[shop, column].dropna().unique().astype(str).tolist())
        return sorted(self.df['Todas', column].dropna().unique().astype(str).tolist())

    def get_products_by_barcode_and_shop(self, barcode, shop):
        """Busca todos os produtos pelo código de barras para a sorveteria atual."""
        try:
            # Filtrar pelo código de barras
            products = self.df[self.df['Todas', 'Codigo de Barras'].str.strip() == barcode.strip()]

            # Filtrar produtos com preço definido na loja atual
            products = products[pd.notna(products[(shop, 'Preco')])]

            if not products.empty:
                return products
            else:
                return pd.DataFrame()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao buscar produtos: {e}")
            return pd.DataFrame()

    def get_products_by_barcode(self, barcode):
        """Retorna todos os produtos com o código de barras especificado em qualquer loja."""
        return self.df[self.df['Todas', 'Codigo de Barras'].str.strip() == barcode.strip()]

class Sale:
    def __init__(self, product_db, shop, payment_method=""):
        self.product_db = product_db
        self.shop = shop
        self.payment_method = payment_method
        self.current_sale = {}
        self.final_price = 0.0
        self.id = str(uuid.uuid4())

    def apply_promotion(self):
        total_price = 0.0
        category_quantities = {}

        # Soma as quantidades por categoria
        for product in self.current_sale.values():
            category = product['categoria']
            quantity = product['quantidade']
            category_quantities[category] = category_quantities.get(category, 0) + quantity

        # Calcula o preço total com promoções
        for product in self.current_sale.values():
            category = product['categoria']
            quantity = product['quantidade']
            promo_qty = product['promo_qt']
            price = product['preco']
            promo_price = product['promo_preco']

            if (self.payment_method in ['Pix', 'Dinheiro'] and
                    promo_qty is not None and
                    category_quantities[category] >= promo_qty):
                total_price += promo_price * quantity
            else:
                total_price += price * quantity

        self.final_price = total_price
        return self.final_price

    def add_product(self, product):
        excel_row = product[('Metadata', 'Excel Row')]
        if excel_row not in self.current_sale or (type(excel_row) == str and excel_row.startswith('Manual')):
            self.current_sale[excel_row] = {
                'categoria': product[('Todas', 'Categoria')],
                'sabor': product[('Todas', 'Sabor')],
                'preco': product[(self.shop, 'Preco')],
                'promo_preco': product.get((self.shop, 'Promo Preco'), product[(self.shop, 'Preco')]),
                'promo_qt': product.get((self.shop, 'Promo Quantidade'), 1),  # int
                'quantidade': 1,
                'indexExcel': excel_row
            }
        else:
            self.current_sale[excel_row]['quantidade'] += 1

    def remove_product(self, excel_row):
        if excel_row in self.current_sale:
            del self.current_sale[excel_row]

    def update_quantity(self, excel_row, quantity):
        if excel_row in self.current_sale:
            self.current_sale[excel_row]['quantidade'] = max(quantity, 0)
            if self.current_sale[excel_row]['quantidade'] == 0:
                self.remove_product(excel_row)

class POSApplication:
    def __init__(self, root):
        self.root = root
        self.root.withdraw()  # Hide the root window initially

        # Ensure Num Lock is always on
        set_numlock(True)
        self.root.bind_all("<Num_Lock>", lambda event: (set_numlock(state=True), "break")[1])

        # Screen scaling
        self.screen_width = self.root.winfo_screenwidth()
        self.screen_height = self.root.winfo_screenheight()
        self.scale_factor = min(self.screen_width / BASE_WIDTH, self.screen_height / BASE_HEIGHT)

        # Initialize product database
        self.product_db = ProductDatabase()

        # Selected shop variable
        self.selected_shop_var = tk.StringVar()

        # Initialize payment method variable here
        self.payment_method_var = tk.StringVar(value="")  # Initialized here

        # Initialize sale
        self.sale = None
        self.stored_sales = []

        # Dictionary to keep track of widgets for each product
        self.product_widgets = {}

        # Initialize UI components
        self.initialize_ui()

        self.manual_add_count = 0
        self.manual_add_list = []

    def initialize_ui(self):
        self.select_shop_window()

    def select_shop_window(self):
        def on_shop_select():
            selected = shop_combobox.get().strip()
            if selected:
                self.selected_shop_var.set(selected)
                shop_window.destroy()
                self.sale = Sale(self.product_db, selected, self.payment_method_var.get())
                self.build_main_window()
            else:
                messagebox.showerror("Erro", "Selecione a loja para continuar.")

        shop_window = tk.Toplevel(self.root)
        shop_window.title("Selecione a loja")
        shop_window.configure(bg="#8b0000")
        shop_window.attributes("-topmost", True)

        # Scale geometry
        win_width = int(400 * self.scale_factor)
        win_height = int(200 * self.scale_factor)
        shop_window.geometry(f"{win_width}x{win_height}")
        shop_window.resizable(False, False)
        shop_window.grab_set()  # Make modal

        # Center the window
        shop_window.update_idletasks()
        x = (shop_window.winfo_screenwidth() // 2) - (win_width // 2)
        y = (shop_window.winfo_screenheight() // 2) - (win_height // 2)
        shop_window.geometry(f"+{x}+{y}")

        # Fonts
        title_font = ("Arial", int(20 * self.scale_factor), "bold")
        version_font = ("Arial", int(8 * self.scale_factor))
        combobox_font = ("Arial", int(14 * self.scale_factor))

        # Label
        tk.Label(shop_window, text="Selecione a loja", bg="#8b0000", fg="#ffffff",
                 font=title_font).pack(pady=int(10 * self.scale_factor))

        tk.Label(shop_window, text=f"Versao: {Version}", bg="#8b0000", fg="#ffffff",
                 font=version_font).pack(pady=0)


        # Combobox
        shop_combobox = ttk.Combobox(
            shop_window,
            values=self.product_db.shops,  # Usando as lojas definidas no ProductDatabase
            state="readonly",
            font=combobox_font,
            width=20
        )
        shop_combobox.pack(pady=int(10 * self.scale_factor))
        shop_combobox.focus()

        # Select Button
        select_btn = ttk.Button(shop_window, text="Selecionar", command=on_shop_select)
        select_btn.pack(pady=int(20 * self.scale_factor))

        self.root.wait_window(shop_window)

    def minimize_application(self):
        """Minimiza a janela principal."""
        self.root.iconify()

    def build_main_window(self):
        self.root.deiconify()
        self.root.title("Sorveteria Lolla")
        self.root.attributes('-fullscreen', True)
        self.root.configure(bg="#1a1a2e")

        # Fonts
        title_font = ("Arial", int(35 * self.scale_factor), "bold")
        shop_font = ("Arial", int(32 * self.scale_factor))
        entry_font = ("Arial", int(16 * self.scale_factor))
        product_list_font = ("Arial", int(16 * self.scale_factor))
        label_font = ("Arial", int(16 * self.scale_factor))
        button_font = ("Arial", int(14 * self.scale_factor))
        final_price_font = ("Arial", int(50 * self.scale_factor), "bold")

        # Title Label
        title_label = ttk.Label(self.root, text="Sorveteria Lolla", font=title_font, background="#1a1a2e",
                                foreground="#ffffff")
        title_label.grid(row=0, column=0, columnspan=3, pady=int(20 * self.scale_factor),
                         padx = int(20 * self.scale_factor), sticky="nw")

        # Configure grid
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=0)
        self.root.grid_columnconfigure(2, weight=1)
        self.root.grid_rowconfigure(2, weight=0)
        self.root.grid_rowconfigure(3, weight=1)

        # Frame para Botões de Controle (Minimizar e Fechar)
        control_frame = tk.Frame(self.root, bg="#1a1a2e")
        control_frame.grid(row=0, column=2, sticky="ne", padx=int(10 * self.scale_factor), pady=int(10 * self.scale_factor))

        # Fechar Botão
        close_button = tk.Button(
            control_frame, text="✖", command=self.close_application, bg="#1a1a2e", fg="red",
            font=("Arial", int(16 * self.scale_factor)), borderwidth=0, width=3
        )
        close_button.pack(side=tk.RIGHT, padx=(0, 5))

        # Minimizar Botão
        minimize_button = tk.Button(
            control_frame, text="—", command=self.minimize_application, bg="#1a1a2e", fg="#ffffff",
            font=("Arial", int(16 * self.scale_factor)), borderwidth=0, width=3
        )
        minimize_button.pack(side=tk.RIGHT)

        # Selected Shop Label
        selected_shop_label = ttk.Label(
            self.root, text=f"{self.selected_shop_var.get()}", background="#1a1a2e",
            foreground="#ffffff", font=shop_font
        )
        selected_shop_label.grid(
            row=0, column=0, columnspan=3, padx=int(20 * self.scale_factor),
            pady=int(70 * self.scale_factor), sticky="sw"
        )

        # Barcode Entry
        self.barcode_entry = ttk.Combobox(self.root, state="normal", font=entry_font, width=45)
        self.barcode_entry.grid(
            row=0, column=0, columnspan=3, padx=int(0 * self.scale_factor),
            pady=int(5 * self.scale_factor), sticky=""
        )
        self.barcode_entry.bind('<Return>', self.handle_barcode)
        self.barcode_entry.bind('<KeyRelease>', self.search_products)

        # Sale Frame
        self.sale_frame = tk.Frame(self.root, bg="#1a1a2e")
        self.sale_frame.grid(
            row=2, column=0, padx=int(10 * self.scale_factor),
            pady=int(10 * self.scale_factor), sticky="nw"
        )

        #Stored Sale Frame
        self.stored_sale_frame = tk.Frame(self.root, bg="#1a1a2e")
        self.stored_sale_frame.grid(
            row=2, column=0, padx=int(30 * self.scale_factor),
            pady=int(700 * self.scale_factor), sticky="nw"
        )

        # Final Price Label
        self.final_price_label = ttk.Label(
            self.root, text="R$0.00", font=final_price_font, background="#1a1a2e",
            foreground="#ffffff"
        )
        self.final_price_label.grid(
            row=1, column=2, columnspan=3, pady=int(25 * self.scale_factor),
            padx=int(50 * self.scale_factor), sticky="ne"
        )

        #Status
        self.status_label = ttk.Label(
            self.root, text="", background="#1a1a2e",
            foreground="#ff8888", font=label_font
        )
        self.status_label.grid(
            row=1, column=2, columnspan=3, padx=int(50 * self.scale_factor),
            pady=int(5 * self.scale_factor), sticky="se"
        )



        valor_pago_label = ttk.Label(
            self.root, text="Valor Pago:", background="#1a1a2e",
            foreground="#ffffff", font=label_font
        )
        valor_pago_label.grid(
            row=2, column=2, padx=int(200 * self.scale_factor),
            pady=int(5 * self.scale_factor), sticky="ne"
        )


        self.valor_pago_entry = ttk.Entry(
            self.root, font=entry_font, width=10
        )
        self.valor_pago_entry.grid(
            row=2, column=2, padx=int(50 * self.scale_factor),
            pady=int(5 * self.scale_factor), sticky="ne"
        )
        self.valor_pago_entry.bind("<KeyRelease>", self.calcular_troco)


        self.troco_label = ttk.Label(
            self.root, text="", background="#1a1a2e",
            foreground="#ffffff", font=label_font
        )
        self.troco_label.grid(
            row=2, column=2, padx=int(50 * self.scale_factor),
            pady=int(40 * self.scale_factor), sticky="ne"
        )



        # Payment Method
        payment_method_label = ttk.Label(
            self.root, text="Método de pagamento:", background="#1a1a2e",
            foreground="#ffffff", font=label_font
        )
        payment_method_label.grid(
            row=2, column=2, padx=int(80 * self.scale_factor),
            pady=int(100 * self.scale_factor), sticky="ne"
        )

        payment_method_combobox = ttk.Combobox(
            self.root, textvariable=self.payment_method_var,
            values=["", "Débito", "Pix", "Dinheiro", "Crédito"],
            state="readonly", font=entry_font, width=20
        )
        payment_method_combobox.grid(
            row=2, column=2, padx=int(50 * self.scale_factor),
            pady=int(130 * self.scale_factor), sticky="ne"
        )
        payment_method_combobox.current(0)
        payment_method_combobox.bind("<<ComboboxSelected>>", self.update_payment_method)


        pay_button = tk.Button(
            self.root, text="Cobrar", font=button_font,
            command=self.cobrar, width=23, height=2
        )
        pay_button.grid(
            row=2, column=2, padx=int(50 * self.scale_factor),
            pady=int(170 * self.scale_factor), sticky="ne"
        )

        #Finalize Sale Button
        finalize_sale_button = tk.Button(
            self.root, text="Finalizar compra", font=button_font,
            command=lambda: self.finalize_sale(self.sale.id), width=23, height=2
        )

        finalize_sale_button.grid(
            row=2, column=2, padx=int(50 * self.scale_factor),
            pady=int(300 * self.scale_factor), sticky="ne"
        )

        # Clear Sale Button
        clear_sale_button = tk.Button(
            self.root, text="Nova venda", font=button_font,
            command=self.new_sale, width=23, height=1
        )
        clear_sale_button.grid(
            row=2, column=2, padx=int(50 * self.scale_factor),
            pady=int(375 * self.scale_factor), sticky="ne"
        )

        # Add New Product Button
        add_new_product_button = tk.Button(
            self.root, text="Cadastrar novo produto", command=self.edit_product,
            font=button_font, width=23, height=1
        )
        add_new_product_button.grid(
            row=2, column=2, padx=int(50 * self.scale_factor),
            pady=int(430 * self.scale_factor), sticky="ne"
        )

        # History button
        history_button = tk.Button(
            self.root, text="Historico", command=self.open_sales_history,
            font=button_font, width=23, height=1
        )
        history_button.grid(
            row=2, column=2, padx=int(50 * self.scale_factor),
            pady=int(485 * self.scale_factor), sticky="ne"
        )

        self.update_sale_display()
        self.root.grid_rowconfigure(4, weight=1)

    def open_sales_history(self):
        SalesHistoryWindow(self.root)

    def search_products(self, event=None, force_search=False):
        search_term = self.barcode_entry.get()

        if not search_term.isdigit() or force_search:
            if search_term:
                search_term = search_term.lower()
                shop = self.selected_shop_var.get()

                if ',' in search_term:
                    search_term = search_term.replace(',', '.')

                # Filter products by barcode, category, flavor, or price
                self.filtered_products = self.product_db.df[
                    self.product_db.df[('Todas', 'Codigo de Barras')].str.contains(search_term, case=False, na=False) |
                    self.product_db.df[('Todas', 'Categoria')].str.contains(search_term, case=False, na=False) |
                    self.product_db.df[('Todas', 'Sabor')].str.contains(search_term, case=False, na=False) |
                    self.product_db.df[(shop, 'Preco')].astype(str).str.contains(search_term, case=False, na=False)
                    ]

                # Populate the combobox with filtered products
                self.barcode_entry['values'] = [
                    f"{product['Todas', 'Codigo de Barras']} - {product['Todas', 'Sabor']} ({product['Todas', 'Categoria']}) - R${product[(shop, 'Preco')]:.2f}".replace(
                        '.', ',')
                    for _, product in self.filtered_products.iterrows()
                ]

                if self.barcode_entry['values']:
                    self.barcode_entry.event_generate(
                        "<<ComboboxSelected>>")  # Trigger selection event if results are found

                # Bind selection event to callback
                self.barcode_entry.bind("<<ComboboxSelected>>", self.handle_product_selection)

    def handle_product_selection(self, event):
        # Get selected product details
        selected_index = self.barcode_entry.current()
        if selected_index != -1:  # Ensure a valid selection
            selected_product = self.filtered_products.iloc[selected_index]
            self.sale.add_product(selected_product)
            self.update_sale_display()
            self.barcode_entry.delete(0, 'end')
            self.barcode_entry['values'] = []

    def confirm_read_error(self, barcode):
        def compare(event=None):
            entered_barcode = barcode_input.get()
            if entered_barcode != "":
                if entered_barcode == barcode:
                    self.edit_product(barcode=entered_barcode)
                else:
                    self.barcode_entry.set(entered_barcode)
                    self.handle_barcode()
                barcode_error_window.destroy()

        barcode_error_window = tk.Toplevel(self.root)
        barcode_error_window.title("Possível erro de leitura")
        barcode_error_window.configure(bg="#8b0000")
        barcode_error_window.attributes("-topmost", True)

        # Scale geometry
        win_width = int(400 * self.scale_factor)
        win_height = int(200 * self.scale_factor)
        barcode_error_window.geometry(f"{win_width}x{win_height}")
        barcode_error_window.resizable(False, False)
        barcode_error_window.grab_set()  # Make modal

        # Center the window
        barcode_error_window.update_idletasks()
        x = (barcode_error_window.winfo_screenwidth() // 2) - (win_width // 2)
        y = (barcode_error_window.winfo_screenheight() // 2) - (win_height // 2)
        barcode_error_window.geometry(f"+{x}+{y}")

        # Fonts
        title_font = ("Arial", int(20 * self.scale_factor), "bold")
        input_font = ("Arial", int(14 * self.scale_factor))

        # Label
        tk.Label(
            barcode_error_window,
            text="Escaneie novamente",
            bg="#8b0000",
            fg="#ffffff",
            font=title_font
        ).pack(pady=int(10 * self.scale_factor))

        # Input
        barcode_input = ttk.Entry(
            barcode_error_window,
            font=input_font,
            width=20
        )
        barcode_input.pack(
            padx=int(20 * self.scale_factor),
            pady=int(10 * self.scale_factor)
        )
        barcode_input.bind("<Return>", compare)
        barcode_input.focus_set()  # Focus on the input field

        # Wait for the window to close
        #self.root.wait_window(barcode_error_window)

    def handle_barcode(self, event=None):

        input_barcode = self.barcode_entry.get().strip()
        barcode = self.barcode_entry.get().strip()
        if not barcode:
            return
        current_shop = self.selected_shop_var.get()

        # Se for digitado um valor
        if ',' in input_barcode or '.' in input_barcode:
            try:
                value = float(input_barcode.replace(",", "."))
                self.manual_add_count = self.manual_add_count + 1
                product = {
                    ('Metadata', 'Excel Row'): 'Manual_'+str(self.manual_add_count),
                    ('Todas', 'Categoria'): 'Não cadastrado',
                    ('Todas', 'Sabor'): '',
                    (current_shop, 'Preco'): value,
                    (current_shop, 'Promo Preco'): None,
                    (current_shop, 'Promo Quantidade'): None
                }


                self.manual_add_list.append(product)
                self.sale.add_product(product)
                self.update_sale_display()
                self.barcode_entry.delete(0, 'end')
                return
            except Exception:
                pass


        # Se foi digitado uma pesquisa
        if not input_barcode.isdigit():
            self.barcode_entry.event_generate('<Down>')
            return

        # Obtém todos os produtos com o mesmo código de barras na loja atual
        matching_products = self.product_db.get_products_by_barcode_and_shop(barcode, current_shop)

        if matching_products.empty:
            # Verifica se existem produtos com o mesmo barcode em outras lojas
            other_products = self.product_db.get_products_by_barcode(barcode)
            if other_products.empty:
                # Nenhum produto encontrado, abrir janela para adicionar novo produto
                self.confirm_read_error(barcode=barcode)
                #self.edit_product(barcode=barcode)
            else:
                # Produtos encontrados em outras lojas, abrir janela para adicionar
                # Encontrar a primeira loja que possui o produto
                prefill_store = None
                for store in self.product_db.shops:
                    preco = other_products.iloc[0][(store, 'Preco')]
                    if pd.notna(preco):
                        prefill_store = store
                        break

                if prefill_store:
                    existing_product = other_products.iloc[0]
                    self.edit_product(index_excel=other_products.iloc[0][('Metadata', 'Excel Row')],
                                      prefill_store=prefill_store)
                else:
                    # Se nenhuma loja possui o Preco definido, abrir sem pré-preenchimento
                    self.edit_product(barcode=barcode)
        else:
            if len(matching_products) == 1:
                # Apenas um produto encontrado na loja atual, adiciona diretamente
                product = matching_products.iloc[0]
                self.sale.add_product(product)
                self.update_sale_display()
            else:
                # Múltiplos produtos encontrados na loja atual, abrir seleção
                self.search_products(force_search=True)
                self.barcode_entry.event_generate('<Down>')
                return

        self.barcode_entry.delete(0, 'end')

    def update_payment_method(self, event=None, method=None):
        if method is not None:
            self.payment_method_var.set(method)
        self.sale.payment_method = self.payment_method_var.get()
        self.sale.apply_promotion()
        self.update_sale_display()
        if self.payment_method_var.get() == "Dinheiro":
            self.valor_pago_entry.focus()

    def create_or_update_product_widget(self, excel_row, details):

        if excel_row not in self.product_widgets:
            row = len(self.product_widgets)

            # Nome e categoria
            text_widget = tk.Text(
                self.sale_frame, height=1, width=35, bg="#1a1a2e", fg="#ffffff",
                font=("Arial", 18), bd=0, highlightthickness=0
            )
            text_widget.grid(row=row, column=0, padx=50, pady=2, sticky="w")
            if details['sabor'] == '':
                text_widget.insert(tk.END, f"{details['categoria']}")
            else:
                text_widget.insert(tk.END, f"{details['categoria']} - {details['sabor']}")
            text_widget.tag_configure("bold", font=("Arial", 18, "bold"))
            text_widget.config(state=tk.DISABLED)

            # Label de preço
            price_label = tk.Label(
                self.sale_frame, text="", bg="#1a1a2e",
                fg="#ffffff", font=("Arial", 18)
            )
            price_label.grid(row=row, column=2, padx=5, pady=2)

            # Entry de quantidade
            quantity_var = tk.StringVar(value=str(details['quantidade']))
            quantity_entry = ttk.Entry(
                self.sale_frame, textvariable=quantity_var, width=5, font=("Arial", 18)
            )
            quantity_entry.grid(row=row, column=1, padx=5, pady=2)
            quantity_entry.bind("<KeyRelease>", lambda event: self.update_quantity_dynamic(excel_row, quantity_var))
            quantity_entry.bind("<FocusIn>", self.select_all_text)

            # Botão de remover
            delete_button = tk.Button(
                self.sale_frame, text="✖",
                command=lambda b=excel_row: self.delete_product(b),
                bg="#1a1a2e", fg="#ffffff", font=("Arial", int(16 * self.scale_factor)),
                borderwidth=0
            )
            delete_button.grid(row=row, column=3, padx=5, pady=2)

            self.product_widgets[excel_row] = None
            if type(excel_row) == str and excel_row.startswith('Manual'):
                self.product_widgets[excel_row] = {
                    'text_widget': text_widget,
                    'price_label': price_label,
                    'quantity_entry': quantity_entry,
                    'quantity_var': quantity_var,
                    'delete_button': delete_button
                }
            else:
                edit_button = tk.Button(
                    self.sale_frame, text="✎",
                    command=lambda b=excel_row: self.edit_product(b),
                    bg="#1a1a2e", fg="#ffffff", font=("Arial", int(16 * self.scale_factor)),
                    borderwidth=0
                )
                edit_button.grid(row=row, column=4, padx=5, pady=2)

                # Armazena os widgets no dicionário
                self.product_widgets[excel_row] = {
                    'text_widget': text_widget,
                    'price_label': price_label,
                    'quantity_entry': quantity_entry,
                    'quantity_var': quantity_var,
                    'delete_button': delete_button,
                    'edit_button': edit_button
                }


        else:
            # Atualiza widgets existentes
            widgets = self.product_widgets[excel_row]
            widgets['text_widget'].config(state=tk.NORMAL)
            widgets['text_widget'].delete("1.0", tk.END)
            if details['sabor'] == '':
                widgets['text_widget'].insert(tk.END, f"{details['categoria']}")
            else:
                widgets['text_widget'].insert(tk.END, f"{details['categoria']} - {details['sabor']}")

            widgets['text_widget'].config(state=tk.DISABLED)

            # Atualiza a quantidade
            widgets['quantity_var'].set(str(details['quantidade']))

        # Atualiza o preço com base na promoção
        widgets = self.product_widgets[excel_row]

        #print(self.sale.payment_method in ['Pix', 'Dinheiro'])
        #print(details['promo_qt'] is not None)
        #print(self.category_quantities.get(details['categoria'], 0) >= details['promo_qt'])

        if (self.sale.payment_method in ['Pix', 'Dinheiro'] and details['promo_qt'] is not None and
                self.category_quantities.get(details['categoria'], 0) >= details['promo_qt']):
            price = details['promo_preco']
            fg_color = "#00ff00"
        else:
            price = details['preco']
            fg_color = "#ffffff"
        widgets['price_label'].config(text=f"R${price:.2f}", fg=fg_color)

    def update_sale_display(self, product=None):
        # Aplica promoções e calcula o preço final
        final_price = self.sale.apply_promotion()
        self.final_price_label.config(text=f"R${final_price:.2f}")

        self.payment_method_var.set(self.sale.payment_method)

        # Calcula quantidades por categoria
        self.category_quantities = {}
        for details in self.sale.current_sale.values():
            category = details['categoria']
            quantity = details['quantidade']
            self.category_quantities[category] = self.category_quantities.get(category, 0) + quantity

        # Atualiza widgets existentes ou cria novos
        for excel_row in self.sale.current_sale.keys():
            #try:

            details = None
            product_series = None

            #if product is not None and excel_row == product[('Metadata', 'Excel Row')]:
            #    product_series = product
            #else:
            if type(excel_row) != str:
                product_series = self.product_db.df.loc[excel_row - 3]  # Ajustar para índice do DataFrame
            else:
                #Quando ha produtos manualmente adicionados
                for product in self.manual_add_list:
                    if excel_row == product[('Metadata', 'Excel Row')]:
                        product_series = product


            details = {
                'categoria': product_series[('Todas', 'Categoria')],
                'sabor': product_series[('Todas', 'Sabor')],
                'preco': float(product_series[(self.sale.shop, 'Preco')]),
                'promo_preco': float(product_series[(self.sale.shop, 'Promo Preco')]) if pd.notna(
                    product_series[(self.sale.shop, 'Promo Preco')]) else float(
                    product_series[(self.sale.shop, 'Preco')]),
                'promo_qt': int(product_series[(self.sale.shop, 'Promo Quantidade')]) if pd.notna(
                    product_series[(self.sale.shop, 'Promo Quantidade')]) else None,
                'quantidade': self.sale.current_sale[excel_row]['quantidade'],
                'indexExcel': excel_row
            }
            self.create_or_update_product_widget(excel_row, details)
            #except Exception as e:
            #    messagebox.showerror("Erro", f"Erro ao atualizar produto {excel_row}: {e}")



        # Restaurar o foco no código de barras
        self.root.bind("<Return>", lambda event: (self.barcode_entry.focus(), "break")[1])
        self.root.bind("<F12>", lambda event: (self.F12_press_handle(), "break")[1])
        self.root.bind("<End>", lambda event: (self.F12_press_handle(), "break")[1])

        self.root.bind("<F5>", lambda event: (self.update_payment_method(method = "Débito"), "break")[1])
        self.root.bind("<F6>", lambda event: (self.update_payment_method(method = "Crédito"), "break")[1])
        self.root.bind("<F7>", lambda event: (self.update_payment_method(method = "Pix"), "break")[1])
        self.root.bind("<F8>", lambda event: (self.update_payment_method(method = "Dinheiro"), "break")[1])
        self.root.bind("<F9>", lambda event: (self.cobrar(), "break")[1])
        self.root.bind("<F10>", lambda event: (self.new_sale(), "break")[1])
        self.root.bind("<F11>", lambda event: (self.finalize_sale(internal_id=self.sale.id), "break")[1])

        # Se um produto foi passado, focar no widget de quantidade correspondente
        if product is not None:
            excel_row = product[('Metadata', 'Excel Row')]
            if excel_row in self.product_widgets:
                self.product_widgets[excel_row]['quantity_entry'].focus_set()

        if self.valor_pago_entry.get():
            self.calcular_troco()

        self.create_or_update_sale_widgets(self.sale.id)

    def F12_press_handle(self):
        self.barcode_entry.focus()
        self.barcode_entry.delete(0, 'end')

    def select_all_text(self, event):
        event.widget.select_range(0, 'end')
        event.widget.icursor('end')
        return 'break'

    def update_quantity_dynamic(self, index_excel, quantity_var):
        try:
            new_quantity = int(quantity_var.get())
            if new_quantity <= 0:
                new_quantity = 0
                self.delete_product(index_excel)
            if index_excel in self.sale.current_sale:
                self.sale.current_sale[index_excel]['quantidade'] = max(new_quantity, 0)
            self.update_sale_display()
        except ValueError:
            if quantity_var.get() != "" :
                messagebox.showerror("Quantidade Inválida", f"Por favor, insira um número válido. ,{quantity_var.get()},")

    def delete_product(self, excel_row):
        self.sale.remove_product(excel_row)
        if excel_row in self.product_widgets:
            # Destroy all widgets associated with the product
            for widget in self.product_widgets[excel_row].values():
                if isinstance(widget, (tk.Widget, ttk.Widget)):
                    widget.grid_forget()
                    widget.destroy()
            # Remove the product from the dictionary
            del self.product_widgets[excel_row]

        # Reindex and reposition remaining widgets
        for idx, (key, widgets) in enumerate(self.product_widgets.items()):
            # Update the grid positions for each widget
            widgets['text_widget'].grid(row=idx, column=0, padx=50, pady=2, sticky="w")
            widgets['quantity_entry'].grid(row=idx, column=1, padx=5, pady=2)
            widgets['price_label'].grid(row=idx, column=2, padx=5, pady=2)
            widgets['delete_button'].grid(row=idx, column=3, padx=5, pady=2)
            if 'edit_button' in widgets:  # Only for non-manual entries
                widgets['edit_button'].grid(row=idx, column=4, padx=5, pady=2)

        self.update_sale_display()

    def edit_product(self, index_excel=None, barcode=None, prefill_store=None):

        shop = self.sale.shop if prefill_store is None else prefill_store
        excel_row = None
        if index_excel is not None:
            #excel_row = self.sale.current_sale[index_excel]['indexExcel']
            excel_row = index_excel
            product_series = self.product_db.df.loc[excel_row - 3]  # Ajustar para índice do DataFrame
            # Obter os dados atuais do produto
            current_barcode = product_series[('Todas', 'Codigo de Barras')]
            current_sabor = product_series[('Todas', 'Sabor')]
            current_categoria = product_series[('Todas', 'Categoria')]
            current_preco = product_series[(shop, 'Preco')]
            current_promo_preco = product_series[(shop, 'Promo Preco')]
            current_promo_qt = product_series[(shop, 'Promo Quantidade')]

        else:
            excel_row = self.product_db.df.index[-1] + 4
            current_barcode = "" if barcode is None else barcode
            current_sabor = ""
            current_categoria = ""
            current_preco = ""
            current_promo_preco = ""
            current_promo_qt = ""



        def save_changes():
            try:
                # Obter e limpar os valores dos campos
                new_barcode = barcode_entry.get().strip()
                new_sabor = sabor_entry.get().strip()
                new_categoria = categoria_entry.get().strip()
                new_preco = preco_entry.get().strip()
                new_promo_preco = promo_preco_entry.get().strip()
                new_promo_qt = promo_qt_entry.get().strip()

                # Validação dos campos obrigatórios
                if not all([new_barcode, new_sabor, new_categoria, new_preco]):
                    messagebox.showerror("Erro", "Preencha todos os campos necessários")
                    return

                # Conversão para valores numéricos, permitindo ',' ou '.'
                new_preco_val = parse_float(new_preco)
                new_promo_preco_val = parse_float(new_promo_preco) if new_promo_preco else None
                new_promo_qt_val = parse_int(new_promo_qt) if new_promo_qt else None

                # Construir o dicionário 'product_info' para atualização
                product_info = {
                    'indexExcel': excel_row,
                    'barcode': new_barcode,
                    'sabor': new_sabor,
                    'categoria': new_categoria,
                    'preco': new_preco_val,
                    'promo_preco': new_promo_preco_val,
                    'promo_qt': new_promo_qt_val,
                }

                # Adicionar ou atualizar o produto na base de dados
                self.product_db.add_product(product_info, self.sale.shop)

                # Atualizar os detalhes na venda atual, se o produto estiver na venda
                if excel_row in self.sale.current_sale:
                    self.sale.current_sale[excel_row].update({
                        'categoria': new_categoria,
                        'sabor': new_sabor,
                        'preco': new_preco_val,
                        'promo_preco': new_promo_preco_val if new_promo_preco_val is not None else new_preco_val,
                        'promo_qt': new_promo_qt_val if new_promo_qt_val is not None else None,
                    })

                    details = {
                        'categoria': new_categoria,
                        'sabor': new_sabor,
                        'preco': new_preco_val,
                        'promo_preco': new_promo_preco_val if new_promo_preco_val is not None else new_preco_val,
                        'promo_qt': new_promo_qt_val if new_promo_qt_val is not None else None,
                        'quantidade': self.sale.current_sale[excel_row]['quantidade'],
                        'indexExcel': excel_row
                    }
                    self.create_or_update_product_widget(excel_row, details)

                self.update_sale_display()  # Atualizar a exibição da venda
                edit_window.destroy()
            except Exception as e:
                messagebox.showerror("Erro", f"Falha ao atualizar o produto: {e}")

        def parse_float(value):
            """Converte um valor para float, aceitando ',' ou '.' como separador decimal."""
            if value:
                try:
                    return float(value.replace(',', '.'))
                except ValueError:
                    raise ValueError(f"Valor inválido para número decimal: {value}")
            return None

        def parse_int(value):
            """Converte um valor para int, aceitando valores que são floats com .0"""
            if value:
                try:
                    # Tenta converter para float primeiro
                    float_val = float(value.strip())
                    if float_val.is_integer():
                        return int(float_val)
                    else:
                        raise ValueError(f"Valor inválido para número inteiro: {value}")
                except ValueError:
                    raise ValueError(f"Valor inválido para número inteiro: {value}")
            return None

        # Criar a janela de edição
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Editar Produto")
        edit_window.configure(bg="#8b0000")
        edit_window.attributes("-topmost", True)

        # Frame para inputs
        input_frame = tk.Frame(edit_window, bg="#8b0000")
        input_frame.pack(pady=10, padx=10)

        # Código de Barras
        tk.Label(input_frame, text="Código de Barras:", bg="#8b0000", fg="#ffffff").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        barcode_entry = tk.Entry(input_frame)
        barcode_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        barcode_entry.insert(0, current_barcode)

        # Sabor
        tk.Label(input_frame, text="Sabor:", bg="#8b0000", fg="#ffffff").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        sabor_entry = ttk.Combobox(
            input_frame,
            values=self.product_db.get_unique_values('Sabor'),
            state="normal"
        )
        sabor_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        sabor_entry.insert(0, current_sabor)

        # Categoria
        tk.Label(input_frame, text="Categoria:", bg="#8b0000", fg="#ffffff").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        categoria_entry = ttk.Combobox(
            input_frame,
            values=self.product_db.get_unique_values('Categoria'),
            state="normal"
        )
        categoria_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        categoria_entry.insert(0, current_categoria)

        # Preço
        tk.Label(input_frame, text="Preço:", bg="#8b0000", fg="#ffffff").grid(row=3, column=0, padx=5, pady=5, sticky="e")
        preco_entry = ttk.Entry(input_frame)
        preco_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        preco_entry.insert(0, current_preco)

        # Promo Preço
        tk.Label(input_frame, text="Promo Preço:", bg="#8b0000", fg="#ffffff").grid(row=4, column=0, padx=5, pady=5, sticky="e")
        promo_preco_entry = ttk.Entry(input_frame)
        promo_preco_entry.grid(row=4, column=1, padx=5, pady=5, sticky="w")
        if pd.notna(current_promo_preco):
            promo_preco_entry.insert(0, current_promo_preco)

        # Promo Quantidade
        tk.Label(input_frame, text="Promo Quantidade:", bg="#8b0000", fg="#ffffff").grid(row=5, column=0, padx=5, pady=5, sticky="e")
        promo_qt_entry = ttk.Entry(input_frame)
        promo_qt_entry.grid(row=5, column=1, padx=5, pady=5, sticky="w")
        if pd.notna(current_promo_qt):
            promo_qt_entry.insert(0, current_promo_qt)

        # Botão para salvar alterações
        save_button = ttk.Button(edit_window, text="Salvar Alterações", command=save_changes)
        save_button.pack(pady=20)

    def calcular_troco(self, event=None):
        try:
            valor_pago = float(self.valor_pago_entry.get().replace(",", "."))
            troco = valor_pago - self.sale.final_price
            if troco < 0.0:
                self.troco_label.config(text=f"Troco: R${troco:.2f}", foreground="#ff5555")
            else:
                self.troco_label.config(text=f"Troco: R${troco:.2f}", foreground="#55ff55")
        except ValueError:
            self.troco_label.config(text="")

    def finalize_sale(self, internal_id):
        def process_sale():  # Define the threaded process
            # Load or create sales history
            try:
                sales_history = pd.read_excel('Historico_vendas.xlsx')
            except FileNotFoundError:
                sales_history = pd.DataFrame(
                    columns=['Data', 'Horario', 'Preco Final', 'Metodo de pagamento', 'Produtos',
                             'Quantidade de produtos']
                )

            # Append new sale
            updated_sales_history = pd.concat([sales_history, sale_df], ignore_index=True)
            updated_sales_history.to_excel('Historico_vendas.xlsx', index=False)

            #print(f"Time taken for finalize_sale: {time.time() - start_time:.2f} seconds")

        start_time = time.time()  # Record the start time
        sale = next((sale for sale in self.stored_sales if sale.id == internal_id), None)

        if not sale.current_sale:
            messagebox.showerror("Erro", "Sem produtos nas vendas!")
            return

        # Apply promotion and calculate final price
        final_price = sale.apply_promotion()

        # Save sale details
        now = datetime.now()
        sale_data = {
            'Data': [now.strftime('%Y-%m-%d')],
            'Horario': [now.strftime('%H:%M:%S')],
            'Preco Final': [final_price],
            'Metodo de pagamento': [sale.payment_method],
            'Produtos': [sale.current_sale],
            'Quantidade de produtos': [sum(product['quantidade'] for product in sale.current_sale.values())]
        }
        sale_df = pd.DataFrame(sale_data)
        threading.Thread(target=process_sale).start()
        self.delete_stored_sale(sale.id)

        #print(f"Time taken for creating new sale: {time.time() - start_time :.2f} seconds")

    def new_sale(self, sale_= None):
        # Reset the sale object
        if sale_ is None:
            self.sale = Sale(self.product_db, self.selected_shop_var.get(), payment_method="")
        else:
            self.sale = sale_

        # Clear all widgets from the sale frame
        for widget in self.sale_frame.winfo_children():
            widget.grid_forget()
            widget.destroy()

        # Clear the product widgets dictionary
        self.product_widgets.clear()

        # Reset other UI elements
        self.payment_method_var.set("")
        self.valor_pago_entry.delete(0, 'end')
        self.troco_label.config(text="")
        self.update_status("")
        self.barcode_entry.delete(0, 'end')

        # Refresh the display
        self.update_sale_display()

    def create_or_update_sale_widgets(self, id):
        if id not in [sale.id for sale in self.stored_sales]:
            self.stored_sales.append(self.sale)

            col = len(self.stored_sales)

            text_button = tk.Button(
                self.stored_sale_frame, text=f"R${self.sale.final_price:.2f}",
                font=("Arial", 18, "bold"), bg="#1a1a2e", fg="#ffffff",
                anchor="w",
                bd=0, highlightthickness=0,
                command=lambda id_=id: self.open_sale(id_)
            )
            text_button.tag = id
            text_button.grid(row=0, column=col, padx=int(60 * self.scale_factor), pady=0, sticky="w")

            # Botão de remover
            delete_button = tk.Button(
                self.stored_sale_frame, text="✖",
                command=lambda id_=id: self.delete_stored_sale(id_),
                bg="#1a1a2e", fg="#ffffff", font=("Arial", int(16 * self.scale_factor)),
                borderwidth=0
            )
            delete_button.tag = id
            delete_button.grid(row=0, column=col, padx=int(18 * self.scale_factor), pady=0, stick="e")


        for widget in self.stored_sale_frame.winfo_children():
            if getattr(widget, "tag", None) == id and widget['text'] != "✖":
                widget['text'] = f"R${self.sale.final_price:.2f}"
                widget['font'] = ("Arial", 22, "bold")
            else:
                widget['font'] = ("Arial", 18)

    def delete_stored_sale(self, id):

        # Find and remove the sale from the stored sales list
        sale_to_remove = next((sale for sale in self.stored_sales if sale.id == id), None)
        if sale_to_remove:
            self.stored_sales.remove(sale_to_remove)

            # Iterate through the widgets in the stored_sale_frame
            for widget in self.stored_sale_frame.winfo_children():
                # Compare the tag of the widget (which stores the sale's ID)
                if hasattr(widget, "tag") and widget.tag == id:
                    widget.grid_forget()  # Remove the widget from the layout
                    widget.destroy()  # Destroy the widget



        for sale in self.stored_sales:
            if sale.final_price == 0.00:
                self.open_sale(sale.id)
                self.rearrange_sale_widgets()
                return

        if self.sale.id == id:
            self.new_sale()
        self.rearrange_sale_widgets()

    def rearrange_sale_widgets(self):
        # Reorganize remaining sale buttons if necessary
        for col, widget in enumerate(self.stored_sale_frame.winfo_children()):
            if widget["text"] != "✖":
                widget.grid_configure(column=col)
            else:
                widget.grid_configure(column=col-1)

    def open_sale(self, id):
        sale = next((sale for sale in self.stored_sales if sale.id == id), None)
        if sale:
            self.new_sale(sale)
            self.update_sale_display()

    def close_application(self):
        self.root.quit()
        self.root.destroy()

    def cobrar(self):
        final_price = self.sale.apply_promotion()
        if final_price >= 1.00:
            self.payment(pay_amount=final_price, payment_type=self.sale.payment_method, internal_id=self.sale.id)

    def create_payment_intent_card(self, amount, internal_id):

        url = "https://api.mercadopago.com/point/integration-api/devices/" + config.device + "/payment-intents"
        headers = {
            "Authorization": "Bearer " + config.id_token,
            "Content-Type": "application/json"
        }
        payload = {
            "amount": amount * 100,
            "description": "Lolla sorveteria",
            "additional_info": {
                "external_reference": internal_id,
                "print_on_terminal": True
            }
        }
        try:
            response = requests.post(url, headers=headers, json=payload)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            return {"error": str(e)}

    def create_payment_intent_debit(self, amount, internal_id):

        url = "https://api.mercadopago.com/point/integration-api/devices/" + config.device + "/payment-intents"
        headers = {
            "Authorization": "Bearer " + config.id_token,
            "Content-Type": "application/json"
        }
        payload = {
            "amount": amount * 100,
            "description": "Lolla sorveteria",
            "payment": {
                "type": "debit_card"
            },
            "additional_info": {
                "external_reference": internal_id,
                "print_on_terminal": True
            }
        }
        try:
            response = requests.post(url, headers=headers, json=payload)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            return {"error": str(e)}

    def create_payment_intent_credit(self, amount, internal_id):

        url = "https://api.mercadopago.com/point/integration-api/devices/" + config.device + "/payment-intents"
        headers = {
            "Authorization": "Bearer " + config.id_token,
            "Content-Type": "application/json"
        }
        payload = {
            "amount": amount * 100,
            "description": "Lolla sorveteria",
            "payment": {
                "installments": 1,
                "installments_cost": "seller",
                "type": "credit_card"
            },
            "additional_info": {
                "external_reference": internal_id,
                "print_on_terminal": True
            }
        }
        try:
            response = requests.post(url, headers=headers, json=payload)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            return {"error": str(e)}

    def create_payment_intent_pix(self, amount, internal_id):
        url = "https://api.mercadopago.com/instore/orders/qr/seller/collectors/" + config.user_id + "/pos/" + config.pos_name + "/qrs"
        headers = {
            "Authorization": "Bearer " + config.id_token,
            "Content-Type": "application/json"
        }
        payload = {
            "external_reference": internal_id,
            "title": "Product order",
            "description": "Purchase description.",
            "total_amount": amount,
            "items": [
                {
                    "title": "Lolla Sorveteria",
                    "unit_price": amount,
                    "quantity": 1,
                    "unit_measure": "unit",
                    "total_amount": amount
                }
            ]
        }
        try:
            response = requests.put(url, headers=headers, json=payload)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            return {"error": str(e)}

    def confirm_payment_card(self, payment_intent_id):

        url = f"https://api.mercadopago.com/point/integration-api/payment-intents/{payment_intent_id}"
        headers = {
            "Authorization": "Bearer " + config.id_token,
        }

        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            return {"error": str(e)}

    def confirm_payment_pix(self):

        url = "https://api.mercadopago.com/instore/qr/seller/collectors/" + config.user_id + "/pos/" + config.pos_name + "/orders"
        headers = {
            "Authorization": "Bearer " + config.id_token,
            "Content-Type": "application/json"
        }

        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            return {"error": str(e)}

    def delete_pix(self):

        url = "https://api.mercadopago.com/instore/qr/seller/collectors/" + config.user_id + "/pos/" + config.pos_name + "/orders"
        headers = {
            "Authorization": "Bearer " + config.id_token,
            "Content-Type": "application/json"
        }

        try:
            response = requests.delete(url, headers=headers)
            response.raise_for_status()
        except requests.exceptions.RequestException as e:
            print(f"error {str(e)}")

    def wait_for_payment_to_finish_card(self, payment_intent_id, internal_id, poll_interval=1):

        while True:
            response = self.confirm_payment_card(payment_intent_id)
            if "state" in response:
                state = response["state"]
                self.update_status(state)

                print(f"Payment state: {state}")
                # if state in ["OPEN", "ON_TERMINAL", "PROCESSING"]:  # Adjust as per terminal state
                if state == "FINISHED":
                    payment_id = response.get("id")
                    print(f"Payment approved with ID: {payment_id}")
                    self.finalize_sale(internal_id)
                    return payment_id

                if state == "CANCELED" or state == "ABANDONED":
                    payment_id = response.get("id")
                    print(state)
                    print(f"Payment canceled/abandoned with ID: {payment_id}")
                    return payment_id
            else:
                print(f"Error checking payment state: {response.get('error', 'Unknown error')}")
            time.sleep(poll_interval)

    def wait_for_payment_to_finish_pix(self, internal_id, qr_window, poll_interval=1):

        while True:
            response = self.confirm_payment_pix()
            # print("response::")
            # print(response)
            if "external_reference" in response:
                external_reference_ = response["external_reference"]
                if internal_id != external_reference_:
                    print(f"{external_reference_} != {internal_id}")
                    return external_reference_
                else:
                    time.sleep(poll_interval)
            else:
                print(f"Payment finished with external_reference {internal_id}")
                qr_window.destroy()  # Close the QR code window
                self.finalize_sale(internal_id)
                return

    def display_qr_code(self, qr_data, internal_id):
        # Generate the QR code
        self.update_status("Gerando QR")

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=int(10 * self.scale_factor),  # Adjust box size
            border=4
        )
        qr.add_data(qr_data)
        qr.make()

        win_width = int(1000 * self.scale_factor)
        win_height = int(1000 * self.scale_factor)
        qr_width = int(900 * self.scale_factor)
        qr_height = int(900 * self.scale_factor)

        # Create an image of the QR code
        img = qr.make_image(fill_color="black", back_color="white")
        img = img.resize((qr_width, qr_height), Image.Resampling.LANCZOS)  # Updated resizing method
        qr_photo = ImageTk.PhotoImage(img)

        # Create a new Tkinter window for the QR code
        qr_window = tk.Toplevel(self.root)
        qr_window.title("Scan QR Code")
        qr_window.configure(bg="#8b0000")
        qr_window.attributes("-topmost", True)

        # Set the window size and make it modal
        qr_window.geometry(f"{win_width}x{win_height}")
        qr_window.resizable(False, False)
        qr_window.grab_set()  # Make the window modal

        # Add the QR code image to the window
        qr_label = tk.Label(qr_window, image=qr_photo, bg="#8b0000")
        qr_label.image = qr_photo  # Keep a reference to avoid garbage collection
        qr_label.pack(expand=True)

        self.update_status("Aguardando pagamento")

        # Wait for payment in a separate thread to prevent UI freezing
        threading.Thread(
            target=self.wait_for_payment_to_finish_pix,
            args=(internal_id, qr_window),
            daemon=True
        ).start()

    def update_status(self, new_status):
        if new_status == "OPEN":
            new_status = "Em aberto"
        if new_status == "FINISHED":
            new_status = "Finalizado"
        if new_status == "ON_TERMINAL":
            new_status = "Na maquininha"
        if new_status == "CANCELED":
            new_status = "Cancelada"
        if new_status == "PROCESSING":
            new_status = "Processando"

        self.status_label.configure(text=f"{new_status}")

    def update_status_thread(self, pay_amount, internal_id):
        # Update the status first before waiting
        self.update_status("Obtendo QR")
        self.delete_pix()
        response = self.create_payment_intent_pix(amount=pay_amount, internal_id=id)

        if "in_store_order_id" in response:
            # After waiting, update status again
            qr = response["qr_data"]
            self.display_qr_code(qr, internal_id)
        else:
            print("Failed to create payment intent.")
            self.update_status("Falha")

    def payment(self, pay_amount, payment_type, internal_id):
        self.update_status("Iniciando pagamento")


        if payment_type == "":
            response = self.create_payment_intent_card(amount=pay_amount, internal_id=internal_id)
            print(response)

            if "id" in response:
                payment_intent_id = response["id"]

                threading.Thread(
                    target=self.wait_for_payment_to_finish_card,
                    args=(payment_intent_id, internal_id,),
                    daemon=True
                ).start()
            else:
                print("Failed to create payment intent.")
                self.update_status("Falha")

        if payment_type == "Débito":
            response = self.create_payment_intent_debit(amount=pay_amount, internal_id=internal_id)
            print(response)

            if "id" in response:
                payment_intent_id = response["id"]

                threading.Thread(
                    target=self.wait_for_payment_to_finish_card,
                    args=(payment_intent_id, internal_id,),
                    daemon=True
                ).start()
            else:
                print("Failed to create payment intent.")
                self.update_status("Falha")

        elif payment_type == "Crédito":
            response = self.create_payment_intent_credit(amount=pay_amount, internal_id=internal_id)
            print(response)

            if "id" in response:
                payment_intent_id = response["id"]
                threading.Thread(
                    target=self.wait_for_payment_to_finish_card,
                    args=(payment_intent_id, internal_id,),
                    daemon=True
                ).start()
            else:
                print("Failed to create payment intent.")
                self.update_status("Falha")

        elif payment_type == "Pix":
            # Run the sleep and update status in a separate thread
            threading.Thread(target=self.update_status_thread, args=(pay_amount, id,), daemon=True).start()

class ToolTip:
    def __init__(self, widget):
        self.widget = widget
        self.tooltip = None
        self.current_item = None  # Track the currently hovered item
        self.widget.bind("<Motion>", self.on_motion)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def on_motion(self, event):
        # Check if the mouse is over a valid item
        item = self.widget.identify_row(event.y)
        if item != self.current_item:  # Only update if the hovered item changes
            self.current_item = item
            if item:
                # Get the 'Produtos' value for the hovered item
                produtos = self.widget.item(item, 'values')[4]
                if produtos:
                    try:
                        # Safely evaluate the string into a dictionary
                        produtos_dict = self.safe_eval_produtos(produtos)

                        # Format the products into a readable list
                        formatted_products = self.format_products(produtos_dict)

                        # Get the position of the mouse
                        bbox = self.widget.bbox(item)
                        if bbox:  # Check if the bounding box is valid
                            x, y, _, _ = bbox
                            x += self.widget.winfo_rootx() + 25
                            y += self.widget.winfo_rooty() + 25

                            # Update or create the tooltip
                            if self.tooltip:
                                # Update the existing tooltip
                                self.tooltip.wm_geometry(f"+{x}+{y}")
                                self.tooltip_label.config(text=formatted_products)
                            else:
                                # Create the tooltip window
                                self.tooltip = tk.Toplevel(self.widget)
                                self.tooltip.wm_overrideredirect(True)
                                self.tooltip.wm_geometry(f"+{x}+{y}")

                                # Add a label to display the formatted products
                                self.tooltip_label = tk.Label(self.tooltip, text=formatted_products, background="#ffffe0", relief="solid", borderwidth=1, justify=tk.LEFT)
                                self.tooltip_label.pack()
                    except Exception as e:
                        print(f"Error formatting products: {e}")
            else:
                # Hide the tooltip if the mouse is not over a valid item
                self.hide_tooltip()

    def hide_tooltip(self, event=None):
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None
            self.current_item = None  # Reset the current item

    def safe_eval_produtos(self, produtos):
        # Define a safe namespace for evaluation
        safe_namespace = {
            "np": np,  # Allow numpy functions
            "int": int,  # Allow Python int
            "float": float,  # Allow Python float
            "nan": np.nan,  # Allow numpy nan
        }

        # Evaluate the string in the safe namespace
        return eval(produtos, {"__builtins__": {}}, safe_namespace)

    def format_products(self, produtos_dict):
        # Format the dictionary into a readable list
        formatted_text = "Produtos:\n"
        for product_id, details in produtos_dict.items():
            # Handle None or nan values in the dictionary
            for key, value in details.items():
                if value is None or (isinstance(value, float) and math.isnan(value)):
                    details[key] = "N/A"

            # Format the price with R$ and two decimal places
            preco = details['preco']
            if isinstance(preco, (int, float)):
                preco = f"R${preco:.2f}"
            else:
                preco = "R$0.00"

            formatted_text += (
                f"- {details['categoria']} ({details['sabor']}): "
                f"{details['quantidade']} unidade(s) a {preco}\n"
            )
        return formatted_text
class SalesHistoryWindow:
    def __init__(self, parent):
        self.parent = parent
        self.window = tk.Toplevel(parent)
        self.window.title("Histórico de Vendas")
        self.window.geometry("800x600")

        # Create a frame to hold the Treeview and scrollbars
        self.tree_frame = tk.Frame(self.window)
        self.tree_frame.pack(fill=tk.BOTH, expand=True)

        # Create a Treeview widget
        self.tree = ttk.Treeview(
            self.tree_frame,
            columns=('Data', 'Horario', 'Preco Final', 'Metodo de pagamento', 'Produtos'),
            show='headings'
        )
        self.tree.heading('Data', text='Data')
        self.tree.heading('Horario', text='Horário')
        self.tree.heading('Preco Final', text='Preço Final')
        self.tree.heading('Metodo de pagamento', text='Método de Pagamento')

        # Hide the 'Produtos' column
        self.tree['displaycolumns'] = ('Data', 'Horario', 'Preco Final', 'Metodo de pagamento')

        # Set column widths
        self.tree.column('Data', width=100, anchor=tk.CENTER)
        self.tree.column('Horario', width=100, anchor=tk.CENTER)
        self.tree.column('Preco Final', width=100, anchor=tk.CENTER)
        self.tree.column('Metodo de pagamento', width=150, anchor=tk.CENTER)

        # Add vertical scrollbar
        self.v_scroll = ttk.Scrollbar(self.tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.v_scroll.set)
        self.v_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree.pack(fill=tk.BOTH, expand=True)

        # Bind hover event to show products
        self.tooltip = ToolTip(self.tree)

        # Load and display sales history
        self.load_sales_history()

    def load_sales_history(self):
        try:
            sales_history = pd.read_excel('Historico_vendas.xlsx')
            # Sort by 'Data' and 'Horario' in descending order to show the latest sales first
            sales_history = sales_history.sort_values(by=['Data', 'Horario'], ascending=[False, False])

            # Replace 'nan' in 'Metodo de pagamento' with an empty string
            sales_history['Metodo de pagamento'] = sales_history['Metodo de pagamento'].replace({np.nan: ""})

            for _, row in sales_history.iterrows():
                # Format the 'Preco Final' with R$ and two decimal places
                preco_final = row['Preco Final']
                if isinstance(preco_final, (int, float)):
                    preco_final = f"R${preco_final:.2f}"
                else:
                    preco_final = "R$0.00"

                self.tree.insert('', 'end', values=(
                    row['Data'],
                    row['Horario'],
                    preco_final,
                    row['Metodo de pagamento'],
                    row['Produtos']
                ))
        except FileNotFoundError:
            messagebox.showerror("Erro", "Arquivo de histórico de vendas não encontrado!")

if __name__ == "__main__":
    root = tk.Tk()
    app = POSApplication(root)
    root.mainloop()
