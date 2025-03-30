from tkinter import ttk, messagebox
import pandas as pd
from openpyxl import load_workbook, Workbook


class ProductDatabase:
    def __init__(self, filepath='Files/produtos.xlsx'):
        self.filepath = filepath
        self.load_products()

    def load_products(self):
        try:
            # Tentar carregar o arquivo com MultiIndex no cabeçalho (2 linhas)
            self.df = pd.read_excel(self.filepath, header=[0, 1], dtype=str)

            # **Validação: Verificar se o DataFrame está vazio**
            if self.df.empty:
                raise ValueError("O arquivo está vazio.")

            # **Validação: Garantir que o cabeçalho tem o formato MultiIndex esperado**
            if not isinstance(self.df.columns, pd.MultiIndex):
                raise ValueError("O arquivo não possui um cabeçalho MultiIndex com duas linhas.")

            # Limpar espaços em branco nos cabeçalhos
            self.df.columns = pd.MultiIndex.from_tuples(
                [(str(x[0]).strip(), str(x[1]).strip()) for x in self.df.columns]
            )

            # Identificar as lojas no nível 0 do MultiIndex (exceto 'Todas')
            self.shops = [shop for shop in self.df.columns.levels[0] if shop != 'Todas']

            # Converter tipos das colunas
            self.df[('Todas', 'Codigo de Barras')] = self.df[('Todas', 'Codigo de Barras')].astype(str)
            self.df[('Metadata', 'Excel Row')] = self.df.index + 3
            for shop in self.shops:
                self.df[(shop, 'Preco')] = pd.to_numeric(self.df[(shop, 'Preco')], errors='coerce')
                self.df[(shop, 'Promo Preco')] = pd.to_numeric(self.df[(shop, 'Promo Preco')], errors='coerce')
                self.df[(shop, 'Promo Quantidade')] = pd.to_numeric(self.df[(shop, 'Promo Quantidade')],
                                                                    errors='coerce')

        except FileNotFoundError:
            messagebox.showerror("Erro", f"Arquivo {self.filepath} não encontrado.")
            self.df = pd.DataFrame()
            self.shops = []  # Inicializa como lista vazia para evitar novos erros
        except ValueError as ve:
            messagebox.showerror("Erro", f"Falha ao carregar produtos: {ve}")
            self.df = pd.DataFrame()
            self.shops = []
        except Exception as e:
            messagebox.showerror("Erro", f"Erro inesperado: {e}")
            self.df = pd.DataFrame()
            self.shops = []

    def add_product(self, product_info, shop):
        try:
            wb = load_workbook(self.filepath)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = "Produtos"
            headers = [
                ("Todas", "Codigo de Barras"),
                ("Todas", "Sabor"),
                ("Todas", "Categoria"),
                (shop, "Preco"),
                (shop, "Promo Preco"),
                (shop, "Promo Quantidade")
            ]
            for col, (header1, header2) in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header1)
                ws.cell(row=2, column=col, value=header2)

        # Map headers to column indices
        header_map = {
            f"{ws.cell(row=1, column=col).value} {ws.cell(row=2, column=col).value}": col
            for col in range(1, ws.max_column + 1)
        }
        excel_row = product_info.get('indexExcel', None)
        if excel_row is None:
            # Encontrar a próxima linha vazia para o código de barras
            data_start_row = 3
            barcode_col = header_map.get("Todas Codigo de Barras")
            excel_row = data_start_row
            while ws.cell(row=excel_row, column=barcode_col).value:
                excel_row += 1

        # Adicionar os dados na mesma linha
        ws.cell(row=excel_row, column=header_map["Todas Codigo de Barras"], value=product_info['barcode'])
        ws.cell(row=excel_row, column=header_map["Todas Sabor"], value=product_info['sabor'])
        ws.cell(row=excel_row, column=header_map["Todas Categoria"], value=product_info['categoria'])

        # Adicionar Preço
        preco = product_info['preco']
        ws.cell(row=excel_row, column=header_map[f"{shop} Preco"], value=float(preco))

        # Adicionar Promo Preço, se disponível
        promo_preco = product_info.get('promo_preco', None)
        ws.cell(row=excel_row, column=header_map[f"{shop} Promo Preco"],
                value=float(promo_preco) if promo_preco is not None else "")

        # Adicionar Promo Quantidade, se disponível
        promo_qt = product_info.get('promo_qt', None)
        ws.cell(row=excel_row, column=header_map[f"{shop} Promo Quantidade"],
                value=int(promo_qt) if promo_qt is not None else "")

        # Salvar e recarregar produtos
        wb.save(self.filepath)
        self.load_products()

    def filter_products(self, search_term, shop):
        if self.df.empty:
            return pd.DataFrame()

        df_copy = self.df.copy()
        df_copy[(shop, 'Preco')] = pd.to_numeric(df_copy[(shop, 'Preco')], errors='coerce')
        df_copy[(shop, 'Preco')] = df_copy[(shop, 'Preco')].apply(
            lambda x: f"{x:.2f}".replace('.', ',') if pd.notnull(x) else "" )

        mask = (
                self.df['Todas', 'Categoria'].str.contains(search_term, case=False, na=False) |
                self.df['Todas', 'Sabor'].str.contains(search_term, case=False, na=False) |
                df_copy[(shop, 'Preco')].astype(str).str.contains(search_term, case=False, na=False)
        )
        return self.df[mask]

    def get_unique_values(self, column, shop=None):
        if shop:
            return sorted(self.df[shop, column].dropna().unique().astype(str).tolist())
        return sorted(self.df['Todas', column].dropna().unique().astype(str).tolist())

    def get_products_by_barcode_and_shop(self, barcode, shop):
        """Busca todos os produtos pelo código de barras para a sorveteria atual."""
        try:
            # Filtrar pelo código de barras
            products = self.df[self.df['Todas', 'Codigo de Barras'].str.strip() == barcode.strip()]

            # Filtrar produtos com preço definido na loja atual
            products = products[pd.notna(products[(shop, 'Preco')])]

            if not products.empty:
                return products
            else:
                return pd.DataFrame()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao buscar produtos: {e}")
            return pd.DataFrame()

    def get_products_by_barcode(self, barcode):
        """Retorna todos os produtos com o código de barras especificado em qualquer loja."""
        return self.df[self.df['Todas', 'Codigo de Barras'].str.strip() == barcode.strip()]
