import openpyxl
import ast
import numpy as np  # To handle NaN values safely

def count_doki_quantities(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    total_quantity = 0

    for row in sheet.iter_rows(min_row=1, max_col=4, max_row=sheet.max_row):
        cell_value = row[3].value  # Column D (index 3)

        # Debugging: Print the raw value
       # print(f"Raw cell value: {repr(cell_value)}")

        if cell_value and isinstance(cell_value, str) and "Doki" in cell_value:
            try:
                # Convert the string into a dictionary (handling 'nan' safely)
                cell_value = cell_value.replace("nan", "None")  # Replace invalid 'nan' with None
                sales_dict = ast.literal_eval(cell_value)  # Convert to dictionary

                # Iterate through all products inside the dictionary
                for sale_data in sales_dict.values():
                    categoria = sale_data.get("categoria", "").lower()
                    quantidade = sale_data.get("quantidade", 0)

                    if "doki" in categoria:
                        total_quantity += quantidade  # Sum all Doki quantities

            except (SyntaxError, ValueError) as e:
                print(f"Skipping invalid row: {cell_value}, Error: {e}")

    return total_quantity

# Run the function
if __name__ == "__main__":
    excel_file = "Files/Historico_vendas.xlsx"
    total = count_doki_quantities(excel_file)
    print(f"Total quantity of 'Doki' products: {total}")
